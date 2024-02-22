import os
import random
import smtplib

from openai import OpenAI
from twilio.rest import Client
from email.message import EmailMessage
from openpyxl import load_workbook
from dotenv import load_dotenv

# Load environment variables from a .env file
load_dotenv()

# Create the client object for the OpenAI API
api_key = os.environ.get('OPENAI_API_KEY')
chat_client = OpenAI(api_key=api_key)

# Load the spreadsheet (Excel file) containing the topics and prompts
workbook = load_workbook('./prompts.xlsx')
sheet = workbook.active

# Get the number of rows in the spreadsheet (excluding the header)
num_rows = sheet.max_row - 1  # Subtract 1 for the header


def tip_generator():
    try:
        # Choose a random row number between 2 and the total number of rows
        random_row = random.randint(2, num_rows + 1)

        # Get the topic and prompt from the randomly selected row
        topic = sheet.cell(row=random_row, column=1).value
        prompt = sheet.cell(row=random_row, column=2).value

        print("Waiting on response...")
        # Use the OpenAI API to generate a response
        response = chat_client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "user", "content": "{}".format(prompt)},
            ]
        )
        print("Response received")

        # Prepare the message to be sent
        message = f"Topic: {topic}\n\n{response.choices[0].message.content}"
        print("Response processed")
        return message
    except Exception as e:
        print(f"Error generating tip: {e}")
        return None


def send_email(body):
    try:
        # Load environment variables from a .env file
        user = os.environ.get('EMAIL_USER')
        password = os.environ.get('EMAIL_PASSWORD')
        to = 'lemmi2102@hotmail.com'
        # Set up the email object
        msg = EmailMessage()

        # Set up the email message
        msg.set_content(body)
        msg['subject'] = "Python tip of the day!\n"
        msg['to'] = to
        msg['from'] = user

        # Set up the SMTP server and send the email
        print("Sending email...")
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(user, password)
        server.send_message(msg)
        server.quit()
        print("Email successfully sent")
    except Exception as e:
        print(f"Error sending email: {e}")


def send_text(body):
    try:
        # Load environment variables from a .env file
        twilio_account_sid = os.environ.get('TWILIO_ACCOUNT_SID')
        twilio_auth_token = os.environ.get('TWILIO_AUTH_TOKEN')
        twilio_phone_number = os.environ.get('TWILIO_PHONE_NUMBER')
        to_phone_number = os.environ.get('PHONE_NUMBER')

        # Initialize Twilio client
        client = Client(twilio_account_sid, twilio_auth_token)
        body = f"Python tip of the day!\n\n{body}"

        # Send a text message
        print("Sending message...")
        client.messages.create(
            body=body,
            from_=twilio_phone_number,
            to=to_phone_number
        )
        print("Message successfully sent")
    except Exception as e:
        print(f"Error sending message: {e}")


if __name__ == '__main__':
    # Generate the Python tip and send it
    tip = tip_generator()
    if tip:
        send_email(tip)
        send_text(tip)
